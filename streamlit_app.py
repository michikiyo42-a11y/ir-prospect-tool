"""
右脳設計 営業リスト自動生成ツール
商材説明を入力するだけで、マッチする企業の
電話番号・部署・担当者名・LinkedIn URLをExcelで出力する
"""

import io
import re
import time
import zipfile
from datetime import datetime, timedelta
from typing import List, Optional, Tuple
from urllib.parse import quote, urljoin

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
EDINET_BASE = "https://disclosure.edinet-fsa.go.jp/api/v2"

# ── キーワード ────────────────────────────────
DX_KEYWORDS = [
    "DX推進", "デジタルトランスフォーメーション", "デジタル変革", "デジタル化",
    "AI活用", "人工知能", "機械学習", "生成AI", "ChatGPT",
    "デジタル人材", "DX人材", "AI人材", "デジタルスキル",
    "リスキリング", "アップスキリング", "人材育成", "研修",
]
PROBLEM_KEYWORDS = [
    "課題", "取り組み中", "推進中", "整備中", "検討中",
    "人材不足", "スキル不足", "育成が必要", "強化が必要",
    "遅れ", "困難", "十分ではない", "今後", "目指",
]
EXEC_TITLES = [
    "CDO", "CIO", "CHRO", "CTO", "最高デジタル責任者", "最高情報責任者",
    "最高人事責任者", "DX推進担当", "デジタル変革推進", "デジタル推進担当",
    "人事部長", "人事本部長", "人材開発部長", "人材育成担当",
    "組織開発部長", "組織開発担当", "研修部長", "研修担当",
    "DX推進部長", "DX推進室長", "HRBP",
]


# ── ページ取得 ────────────────────────────────

def fetch(url: str) -> Optional[BeautifulSoup]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        return BeautifulSoup(resp.content, "html.parser")
    except Exception:
        return None


# ── 商材キーワード抽出 ─────────────────────────

def extract_product_keywords(text: str) -> List[str]:
    """
    商材説明文から検索キーワードを抽出する。
    """
    # よく使われる業界ワードを抽出
    all_kw = DX_KEYWORDS + PROBLEM_KEYWORDS + [
        "人事", "組織", "育成", "研修", "スキル", "学習", "eラーニング",
        "DX", "AI", "デジタル", "変革", "人材",
    ]
    found = [kw for kw in all_kw if kw in text]
    return found[:10] if found else ["DX推進", "人材育成", "AI活用"]


def scrape_product_url(url: str) -> str:
    """商材URLからテキストを取得する。"""
    soup = fetch(url)
    if not soup:
        return ""
    for tag in soup(["script", "style", "nav", "header", "footer"]):
        tag.decompose()
    return "\n".join(l for l in soup.get_text(separator="\n").splitlines() if l.strip())[:3000]


# ── 電話番号スクレイピング ───────────────────────

PHONE_PATTERN = re.compile(r'(?:0\d{1,4}[-－ー]\d{1,4}[-－ー]\d{4}|\d{10,11})')


def find_phone_number(company_name: str, company_website: str = "") -> str:
    """
    会社名から電話番号を探す。
    1. 公式サイトのお問い合わせ/会社概要ページ
    2. Google検索結果（スニペット）
    """
    # 会社サイトのURLが分かっている場合
    if company_website:
        for path in ["", "/contact", "/company", "/about", "/corporate"]:
            soup = fetch(company_website.rstrip("/") + path)
            if soup:
                text = soup.get_text()
                phones = PHONE_PATTERN.findall(text)
                phones = [p for p in phones if len(p.replace("-","").replace("－","")) >= 10]
                if phones:
                    return phones[0]

    # Google検索でスニペットから電話番号を探す
    try:
        query = quote(f"{company_name} 電話番号 代表")
        resp = requests.get(
            f"https://www.google.com/search?q={query}",
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=10
        )
        soup = BeautifulSoup(resp.content, "html.parser")
        text = soup.get_text()
        phones = PHONE_PATTERN.findall(text)
        phones = [p for p in phones if len(p.replace("-","").replace("－","").replace("ー","")) >= 10]
        if phones:
            return phones[0]
    except Exception:
        pass

    return "要確認"


def find_company_website(company_name: str) -> str:
    """会社名から公式サイトURLを探す。"""
    try:
        query = quote(f"{company_name} 公式サイト")
        resp = requests.get(
            f"https://www.google.com/search?q={query}",
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=10
        )
        soup = BeautifulSoup(resp.content, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("/url?q="):
                url = href[7:].split("&")[0]
                # 日本企業の公式サイトパターン
                if any(d in url for d in [".co.jp", ".com", ".jp"]):
                    if not any(x in url for x in ["google", "wikipedia", "amazon", "linkedin"]):
                        return url
    except Exception:
        pass
    return ""


# ── EDINET 検索 ───────────────────────────────

def find_annual_report(company_name: str, log_fn) -> Optional[dict]:
    today = datetime.now()

    def weekdays(start: int, end: int):
        for d in range(start, end):
            dt = today - timedelta(days=d)
            if dt.weekday() < 5:
                yield dt.strftime("%Y-%m-%d")

    ranges = (
        list(weekdays(0, 30)) +
        list(weekdays(300, 430)) +
        list(weekdays(30, 300)) +
        list(weekdays(430, 500))
    )

    for date_str in ranges:
        try:
            resp = requests.get(
                f"{EDINET_BASE}/documents.json",
                params={"date": date_str, "type": 2},
                timeout=10,
            )
            if resp.status_code != 200:
                time.sleep(0.2)
                continue
            for doc in resp.json().get("results", []):
                if (company_name in doc.get("filerName", "")
                        and doc.get("docTypeCode") == "120"):
                    log_fn(f"発見: {doc['filerName']} ({date_str})")
                    return doc
        except Exception:
            pass
        time.sleep(0.15)
    return None


def download_ir_text(doc_id: str) -> str:
    try:
        resp = requests.get(
            f"{EDINET_BASE}/documents/{doc_id}",
            params={"type": 5}, timeout=60,
        )
        resp.raise_for_status()
        texts = []
        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            htm_files = sorted(
                [f for f in zf.namelist() if f.lower().endswith((".htm", ".xhtml"))],
                key=lambda f: zf.getinfo(f).file_size, reverse=True,
            )
            for fname in htm_files[:5]:
                try:
                    soup = BeautifulSoup(zf.read(fname), "html.parser")
                    for tag in soup(["script", "style"]):
                        tag.decompose()
                    text = "\n".join(
                        l for l in soup.get_text(separator="\n").splitlines() if l.strip()
                    )
                    if len(text) > 300:
                        texts.append(text)
                except Exception:
                    continue
        return "\n".join(texts)[:60000]
    except Exception:
        return ""


# ── 導入事例スクレイピング ──────────────────────

def get_case_study_urls(index_url: str) -> List[str]:
    soup = fetch(index_url)
    if not soup:
        return []
    seen, urls = set(), []
    for a in soup.find_all("a", href=True):
        full = urljoin(index_url, a["href"])
        if (full not in seen and full != index_url
                and full.endswith(".html")
                and index_url.split("/")[2] in full):
            seen.add(full)
            urls.append(full)
    return urls


def scrape_case_study(url: str) -> List[dict]:
    soup = fetch(url)
    if not soup:
        return []

    for tag in soup(["script", "style", "nav", "header", "footer"]):
        tag.decompose()
    text = "\n".join(l for l in soup.get_text(separator="\n").splitlines() if l.strip())

    # 会社名
    company = ""
    for pat in [r'株式会社[^\s　、。「」（）]{2,15}',
                r'[^\s　、。「」（）]{2,15}株式会社',
                r'[^\s　、。「」（）]{2,15}ホールディングス']:
        m = re.search(pat, text[:500])
        if m and 4 <= len(m.group()) <= 20:
            company = m.group()
            break

    if not company:
        return []

    # 担当者抽出
    persons = []
    seen_names: set = set()
    photo_pat = re.compile(r'[（(][^）)]*?(左|右|中央|番目)[^）)]*?[）)]')
    head_bracket = re.compile(r'^[（(][^）)]+[）)]\s*')
    tail_name = re.compile(r'([一-鿿]{2,3}[一-鿿]{1,4})\s*$')

    for line in text.splitlines():
        if 'さん' not in line:
            continue
        for seg in line.split('さん')[:-1]:
            clean = photo_pat.sub('', seg).strip()
            clean = head_bracket.sub('', clean).strip()
            m = tail_name.search(clean)
            if not m:
                continue
            name = m.group(1).strip()
            if name in seen_names or len(name) < 3:
                continue
            before = clean[:m.start()].strip().rstrip('　').rstrip()
            parts = [p.strip() for p in re.split(r'[　\s]+', before) if p.strip()]
            dept_role = '　'.join(parts)
            seen_names.add(name)
            persons.append({"name": name, "dept_role": dept_role})

    if not persons:
        return []

    rows = []
    for p in persons[:8]:
        q = quote(f"{p['name']} {company}")
        rows.append({
            "会社名":     company,
            "部署・役職": p["dept_role"],
            "担当者名":   p["name"],
            "ソース":     "導入事例",
            "_url":       url,
        })
    return rows


# ── テキスト分析 ──────────────────────────────

def extract_persons_from_ir(text: str, company: str) -> List[dict]:
    persons = []
    seen: set = set()
    photo_pat = re.compile(r'[（(][^）)]*?(左|右|中央|番目)[^）)]*?[）)]')
    head_bracket = re.compile(r'^[（(][^）)]+[）)]\s*')
    tail_name = re.compile(r'([一-鿿]{2,3}[一-鿿]{1,4})\s*$')

    for title in EXEC_TITLES:
        for m in re.finditer(re.escape(title), text):
            ctx = text[max(0, m.start()-100): min(len(text), m.end()+100)]
            for line in ctx.splitlines():
                if 'さん' not in line and title not in line:
                    continue
                # さん形式
                if 'さん' in line:
                    for seg in line.split('さん')[:-1]:
                        clean = photo_pat.sub('', seg).strip()
                        clean = head_bracket.sub('', clean).strip()
                        nm = tail_name.search(clean)
                        if nm:
                            name = nm.group(1).strip()
                            if len(name) >= 3 and name not in seen:
                                seen.add(name)
                                persons.append({"name": name, "dept_role": title})
                # 役職名前形式
                nm2 = re.search(r'[一-鿿]{2,3}[一-鿿]{1,4}', ctx)
                if nm2:
                    name = nm2.group().strip()
                    if len(name) >= 3 and name not in seen:
                        seen.add(name)
                        persons.append({"name": name, "dept_role": title})
                break
    return persons[:5]


def score_company(text: str, product_keywords: List[str]) -> int:
    kw_hits = sum(text.count(k) for k in product_keywords)
    prob_hits = sum(text.count(k) for k in PROBLEM_KEYWORDS)
    return min(5, max(1, int((kw_hits * 0.3 + prob_hits * 0.5) / 5))) if kw_hits else 1


# ── Excel 出力 ────────────────────────────────

COLUMNS = [
    ("会社名",           22),
    ("電話番号",         18),
    ("部署・役職",       38),
    ("担当者名",         16),
    ("LinkedIn URL",     38),
    ("マッチスコア",     12),
    ("ソース",           12),
]
SCORE_COLORS = {5: "C00000", 4: "FF6600", 3: "FFD700", 2: "92D050", 1: "BFBFBF"}


def build_excel(rows: List[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "営業リスト"

    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    for col, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28

    score_col = next(i+1 for i, (n, _) in enumerate(COLUMNS) if n == "マッチスコア")
    li_col    = next(i+1 for i, (n, _) in enumerate(COLUMNS) if n == "LinkedIn URL")

    for row, r in enumerate(sorted(rows, key=lambda x: x.get("マッチスコア", 0), reverse=True), 2):
        score = r.get("マッチスコア", 1)
        for col, (key, _) in enumerate(COLUMNS, 1):
            c = ws.cell(row=row, column=col, value=r.get(key, ""))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == score_col:
                color = SCORE_COLORS.get(score, "BFBFBF")
                c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                c.font = Font(bold=True, color="FFFFFF" if score >= 4 else "000000", size=13)
                c.alignment = Alignment(horizontal="center", vertical="center")
            if col == li_col:
                c.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[row].height = 45

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 1社まとめて処理 ───────────────────────────

def process_company_ir(company_name: str, product_keywords: List[str], log_fn) -> List[dict]:
    doc_info = find_annual_report(company_name, log_fn)
    if not doc_info:
        return []

    text = download_ir_text(doc_info["docID"])
    filer = doc_info.get("filerName", company_name)
    score = score_company(text, product_keywords)
    persons = extract_persons_from_ir(text, filer)

    log_fn(f"電話番号を検索中...")
    website = find_company_website(filer)
    phone = find_phone_number(filer, website)

    rows = []
    if persons:
        for p in persons:
            q = quote(f"{p['name']} {filer}")
            rows.append({
                "会社名":        filer,
                "電話番号":      phone,
                "部署・役職":    p["dept_role"],
                "担当者名":      p["name"],
                "LinkedIn URL":  f"https://www.linkedin.com/search/results/people/?keywords={q}",
                "マッチスコア":  score,
                "ソース":        "IR文書",
            })
    else:
        # 担当者不明でも会社情報だけ残す
        q = quote(f"DX推進 人事部長 {filer}")
        rows.append({
            "会社名":        filer,
            "電話番号":      phone,
            "部署・役職":    "要確認（DX推進部・人事部）",
            "担当者名":      "（IR記載なし）",
            "LinkedIn URL":  f"https://www.linkedin.com/search/results/people/?keywords={q}",
            "マッチスコア":  score,
            "ソース":        "IR文書",
        })
    return rows


# ── Streamlit UI ──────────────────────────────

st.set_page_config(
    page_title="営業リスト自動生成 | 右脳設計",
    page_icon="📋",
    layout="wide",
)

st.title("📋 営業リスト自動生成ツール")
st.caption("商材を入力するだけで、マッチする企業の電話番号・部署・担当者・LinkedInをExcel出力します")

# ── サイドバー：商材設定 ──────────────────────
with st.sidebar:
    st.header("① 商材を登録する")
    product_mode = st.radio("入力方法", ["テキストで入力", "URLで入力"])

    if product_mode == "テキストで入力":
        product_text = st.text_area(
            "商材・サービスの説明",
            height=180,
            placeholder="例：AI・DX人材の育成を支援する研修ツール。\n主に大手企業の人事部・DX推進部向け。\nリスキリングや生成AI活用研修が中心。",
        )
    else:
        product_url = st.text_input("商材サイトURL", placeholder="https://...")
        product_text = ""
        if product_url and st.button("URLから読み込む"):
            with st.spinner("サイトを読み込み中..."):
                product_text = scrape_product_url(product_url)
            if product_text:
                st.success(f"{len(product_text)}文字読み込みました")
            else:
                st.error("読み込めませんでした")

    product_keywords = []
    if product_text:
        product_keywords = extract_product_keywords(product_text)
        st.markdown("**抽出されたキーワード：**")
        st.write("、".join(product_keywords))

    st.divider()
    st.markdown("""
**スコアの見方**
- 🔴 5：最優先
- 🟠 4：優先
- 🟡 3：検討
- 🟢 2：様子見
""")

# ── メインエリア ──────────────────────────────
tab1, tab2 = st.tabs(["🏢 企業リストから探す", "📰 導入事例記事から探す"])

# ────────────────────────────────────────
# Tab1: EDINET（企業リスト）
# ────────────────────────────────────────
with tab1:
    st.markdown("### 企業名を入力 → 電話番号・担当者を自動取得")

    col1, col2 = st.columns([1, 1])
    with col1:
        company_input = st.text_area(
            "企業名（1行に1社）",
            height=250,
            placeholder="富士通株式会社\n日本電気株式会社\n株式会社日立製作所",
        )
    with col2:
        companies = [c.strip() for c in company_input.splitlines() if c.strip()]
        if companies:
            st.markdown("#### 対象企業")
            for c in companies:
                st.write(f"・{c}")
            st.info(f"{len(companies)}社　推定{len(companies)*3}〜{len(companies)*5}分")

    if st.button("🚀 リスト生成スタート", type="primary", disabled=not companies or not product_keywords):
        if not product_keywords:
            st.warning("先にサイドバーで商材を入力してください")
        else:
            all_rows = []
            progress = st.progress(0)
            status = st.empty()
            log_box = st.empty()

            for i, company in enumerate(companies):
                status.info(f"**[{i+1}/{len(companies)}] {company}** 処理中...")

                def log(msg): log_box.caption(msg)

                rows = process_company_ir(company, product_keywords, log)
                all_rows.extend(rows)
                progress.progress((i+1)/len(companies))
                time.sleep(0.5)

            status.success(f"✅ 完了！{len(all_rows)}件のリストを生成しました")
            log_box.empty()

            if all_rows:
                df = pd.DataFrame(all_rows)[["会社名","電話番号","部署・役職","担当者名","マッチスコア"]]
                st.dataframe(df.sort_values("マッチスコア", ascending=False), use_container_width=True)

                excel = build_excel(all_rows)
                st.download_button(
                    "📥 Excelダウンロード",
                    data=excel,
                    file_name=f"sales_list_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )

# ────────────────────────────────────────
# Tab2: 導入事例記事
# ────────────────────────────────────────
with tab2:
    st.markdown("### 導入事例サイトのURLを入れると担当者を自動抽出")

    col1, col2 = st.columns([1, 1])
    with col1:
        case_urls_input = st.text_area(
            "導入事例の一覧ページURL（1行に1URL）",
            height=200,
            placeholder="https://www.benesse.co.jp/udemy/business/case/\nhttps://...",
        )
        max_articles = st.slider("記事の上限数", 5, 50, 20)

    with col2:
        case_urls = [u.strip() for u in case_urls_input.splitlines() if u.strip()]
        if case_urls:
            st.markdown("#### 入力URL")
            for u in case_urls:
                st.write(f"・{u[:60]}...")

    if st.button("🚀 事例から担当者を抽出", type="primary", disabled=not case_urls):
        all_rows = []
        progress = st.progress(0)
        status = st.empty()

        for ui, index_url in enumerate(case_urls):
            status.info(f"一覧ページを取得中: {index_url[:60]}...")
            article_urls = get_case_study_urls(index_url)
            article_urls = article_urls[:max_articles]

            for i, url in enumerate(article_urls):
                status.info(f"[{i+1}/{len(article_urls)}] 記事を解析中...")
                rows = scrape_case_study(url)

                for r in rows:
                    company = r["会社名"]
                    # 電話番号を検索
                    website = find_company_website(company)
                    phone = find_phone_number(company, website)
                    q = quote(f"{r['担当者名']} {company}")
                    all_rows.append({
                        "会社名":        company,
                        "電話番号":      phone,
                        "部署・役職":    r["部署・役職"],
                        "担当者名":      r["担当者名"],
                        "LinkedIn URL":  f"https://www.linkedin.com/search/results/people/?keywords={q}",
                        "マッチスコア":  3,
                        "ソース":        "導入事例",
                    })

                progress.progress((i+1)/len(article_urls))
                time.sleep(1)

        status.success(f"✅ {len(all_rows)}件を抽出しました")

        if all_rows:
            df = pd.DataFrame(all_rows)[["会社名","電話番号","部署・役職","担当者名"]]
            st.dataframe(df, use_container_width=True)

            excel = build_excel(all_rows)
            st.download_button(
                "📥 Excelダウンロード",
                data=excel,
                file_name=f"case_study_list_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
